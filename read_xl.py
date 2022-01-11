from openpyxl import Workbook
out_wb = Workbook()
out_ws = out_wb.active



from openpyxl import load_workbook
wb = load_workbook(filename = 'doc/asset.xlsx')
field=['物品編號', '年度', '憑單編號', '傳票號碼', '物品', '牌子及型號', 'S/N (P/N)', '數量', '單價(M$) ', '總值(M$)','攤折淨值(M$)', '存放地點', '資助', '資助項目名稱', '備註', '供應商', '登賬日期']
ignore_ws=['資產分類表','工作表1']
ignore_col=['異動   原因','異動原因','*    攤折完成',  '*  攤折完', '* 攤折完成','*  攤折完成','^ 投保產物', '投保項備註','學校資金']

out_ws.append(field)
for sn_ in wb.sheetnames:
    if sn_ in ignore_ws:continue
    #print(sn_)
    ws=wb[sn_]
    col_s=[]
    col_n=[]
    f_col_=None
    for cidx in range(65,90):
        c_name=ws[f"{chr(cidx)}{6}"].value
        if c_name==None: break
        if c_name in ignore_col: continue
        if c_name=="年度":f_col_=chr(cidx)
        if f_col_==None and c_name=="憑單編號":f_col_=chr(cidx)
        col_s.append(c_name)
        col_n.append(chr(cidx))
    print(col_s)
    #print(col_n)
    for ridx in range(7,750):
        f_v=ws[f"{f_col_}{ridx}"].value
        if f_v == None or f_v=="": break
        row_=['', '', '', '', '', '', '', '', '', '','', '', '', '', '', '', '',sn_]
        for idx,v_ in enumerate(col_s):
            cell_v=ws[f"{col_n[idx]}{ridx}"].value
            row_[field.index(v_)]=cell_v
        out_ws.append(row_)



wb.save('doc/out.xlsx')